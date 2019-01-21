#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng


import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from lxml import etree

fenlei = ['无码',
          '欧美',
          '有码']

wb = load_workbook('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
# 获得所有sheet的名称
# print(wb.get_sheet_names())
# 根据sheet名字获得sheet

def download_mag(fenlei):
    for leibie in fenlei:
        a_sheet = wb.get_sheet_by_name(leibie)
        max_row = a_sheet.max_row
        print(max_row)
        for i in range(2, max_row+1):
            # print(a_sheet['G%d' % i].value)
            item = a_sheet['A%d' % i].value
            if a_sheet['G%d' % i].value is not None:
                pass
            else:
                url = a_sheet['E%d' % i].value
                # print(url)

                req = requests.get(url)
                # time.sleep(1)
                html = req.content.decode('utf-8', 'ignore')
                # print(html)
                page = etree.HTML(html)
                # print(page)
                pic_link = page.xpath("/html/body/div[5]/div[1]/div[1]/a/img/@src")
                pic_link2 = ''
                if not pic_link:
                    pic_link2 = '暂时没有图片'
                else:
                    for t in pic_link:
                        pic_link2 = pic_link2 + t + ','
                length = page.xpath("/html/body/div[5]/div[1]/div[2]/p/span[@class='header'][contains(text(), '長度:')]/parent::p/text()")
                length2 = ''
                if not length:
                    length2 = '暂时没有时长数据'
                else:
                    for t in length:
                        length2 = length2 + t + ','

                fenlei = page.xpath("/html/body/div[5]/div[1]/div[2]/p[@class='star-show']/preceding-sibling::p/span[@class='genre']/a/text()")
                # print(fenlei)
                fenlei2 = ''
                if not fenlei:
                    fenlei2 = '暂时没有分类信息'
                else:
                    for t in fenlei:
                        fenlei2 = fenlei2 + t + ','
                actor = page.xpath("/html/body/div[5]/div[1]/div[2]/p[@class='star-show']/following-sibling::p/span[@class='genre']/a/text()")
                # print(actor)
                actor2 = ''
                if not actor:
                    actor2 = '暂时没有演员信息'
                else:
                    for t in actor:
                        actor2 = actor2 + t + ','
                movie_id = re.search(r'\d{11}|\d{10}', html).group()

                url2 = 'https://www.xwb2uooia5o59avwxb.com/ajax/uncledatoolsbyajax.php?gid=%s&lang=zh&uc=1' % movie_id
                if leibie == '欧美':
                    url2 = 'https://www.javbus.work/ajax/uncledatoolsbyajax.php?gid=%s&lang=zh&uc=1' % movie_id
                if leibie == '有码':
                    url2 = 'https://www.xwb2uooia5o59avwxb.com/ajax/uncledatoolsbyajax.php?gid=%s&lang=zh&uc=0' % movie_id
                # print(url2)


                headers = {
                    'cookie':'__cfduid=d761cee020cec276f507132139addc7421545555734; __guid=35400359.4113806376697964000.1545555738264.177; HstCfa4127494=1545555756697; HstCmu4127494=1545555756697; 4fJN_2132_saltkey=HJJ2dBHW; 4fJN_2132_lastvisit=1546156644; existmag=all; PHPSESSID=1geleti43uoi8rs80mv2hhh4o7; HstCnv4127494=4; HstCns4127494=13; HstCla4127494=1546433773520; HstPn4127494=2; HstPt4127494=154; starinfo=glyphicon%20glyphicon-plus; monitor_count=3',
                    'referer':'https://www.xwb2uooia5o59avwxb.com/010918_205',
                    'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
                    'x-requested-with':'XMLHttpRequest'

                }

                file = requests.get(url2, headers=headers)
                # print(file)
                html = file.content.decode('utf-8', 'ignore')
                # print(html)
                page = etree.HTML(html)
                maglist = page.xpath('//td//a[@title="滑鼠右鍵點擊並選擇【複製連結網址】"]/@href')
                # print(len(maglist))
                maglinks = []
                maglinks2 = ''
                for n in range(1,len(maglist),3):
                    maglinks.append(maglist[n-1])
                if not maglinks:
                    maglinks2 = '暂时没有磁力链接'
                else:
                    for t in maglinks:
                        maglinks2 = maglinks2 + t + ','
                print(item)
                print(pic_link2)
                print(length2)
                print(fenlei2)
                print(actor2)
                print(movie_id)
                print(maglinks2)
                a_sheet['F%d' % i] = pic_link2
                a_sheet['G%d' % i] = length2
                a_sheet['H%d' % i] = fenlei2
                a_sheet['I%d' % i] = actor2
                a_sheet['J%d' % i] = movie_id
                a_sheet['K%d' % i] = maglinks2
                print('已获取分类下第%d条信息' % (i-1))
                if i % 10 == 0:
                    wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
        wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')

def trytoget(fenlei):
    try:
        download_mag(fenlei)
    except EnvironmentError as e:
        print(e)
        errorlist.append(e)
        trytoget(fenlei)


errorlist = []
trytoget(fenlei)
print(errorlist)
