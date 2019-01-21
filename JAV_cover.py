#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng


import re
import requests
from openpyxl import load_workbook
import os
from lxml import etree

fenlei = ['有码','无码','欧美']

wb = load_workbook('I:\\JAV全站资源\\新建工作表.xlsx')
# 获得所有sheet的名称
# print(wb.get_sheet_names())
# 根据sheet名字获得sheet

def download_mag(fenlei):

    for leibie in fenlei:
        a_sheet = wb.get_sheet_by_name(leibie)
        max_row = a_sheet.max_row
        print(max_row)
        dirname = 'I:\\JAV全站资源\\JAV封面\\%s\\' % leibie
        if not os.path.exists(dirname):
            os.makedirs(dirname)
        for i in range(2, max_row+1):
            piclink = a_sheet['F%d' % i].value
            print(piclink)
            filename = a_sheet['A%d' % i].value + '.jpg'
            if os.path.exists(dirname + filename):
                pass
            else:
                with open(dirname + filename, 'wb') as f:
                    req = requests.get(piclink).content
                    # html = req.decode('utf-8', 'ignore')
                    f.write(req)
                    print('已下载本分类第%d张封面图:%s' % (i-1, filename))


def trytodownload_mag(fenlei):
    try:
        download_mag(fenlei)
    except EnvironmentError as e:
            print(e)
            errorlist.append(e)
            trytodownload_mag(fenlei)


errorlist = []
trytodownload_mag(fenlei)
print(errorlist)
