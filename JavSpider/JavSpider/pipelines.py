# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
from JavSpider.spiders.youma import YoumaSpider
from JavSpider.spiders.wuma import WumaSpider
from openpyxl import load_workbook
from JavSpider.spiders.oumei import OumeiSpider

# class JavspiderPipeline(object):
#     def process_item(self, item, spider):
#         print(item)
#
#         return item


class JavspiderPipeline(object):
    def __init__(self):
        '''
        initialize the object
        '''
        self.wuma_count = 0
        self.youma_count = 0
        self.oumei_count = 0
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['番号', '标题', '发行日期', '详细页地址', '封面地址', '分类', '演员', '磁力链接'])
    # def open_spider(self, spider):
    #     self.wb = Workbook('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
    #     print(self.wb)
    #     self.ws = self.wb.create_sheet('无码')
    #     # self.ws = self.wb.get_sheet_by_name('无码')
    #     print(self.ws)
    #     self.ws.append(['番号', '标题', '发行日期', '详细页地址', '封面地址', '分类', '演员', '磁力链接'])
        # if spider.name =='youma':
        #
        #     '''
        #     create a queue
        #     :return:
        #     '''
        #     self.wb = openpyxl.Workbook('C:\\Users\\lushe\\Desktop\\有码.xlsx')
        #     self.ws = self.wb.create_sheet('有码')
        #     self.ws.append(['番号', '标题','发行日期','详细页地址','封面地址','分类','演员', '磁力链接'])
        # elif spider.name =='wuma':
        #     self.wb = openpyxl.Workbook('C:\\Users\\lushe\\Desktop\\无码.xlsx')
        #     self.ws = self.wb.create_sheet('无码')
        #     self.ws.append(['番号', '标题','发行日期','详细页地址','封面地址','分类','演员', '磁力链接'])
        # elif spider.name == 'oumei':
        #     self.wb = openpyxl.Workbook('C:\\Users\\lushe\\Desktop\\欧美.xlsx')
        #     self.ws = self.wb.create_sheet('欧美')
        #     self.ws.append(['番号', '标题','发行日期','详细页地址','封面地址','分类','演员', '磁力链接'])

    def process_item(self, item, spider):
        '''
        save every
        :return:
        '''
        if spider.name == 'wuma':
            line = [item['number'],item['title'],item['pdate'],item['detail_link'],item['cover'],item['fenlei'],item['actor'],item['maglinks']]
            self.ws.append(line)
            self.wuma_count += 1
            print('已下载 %s, %d:' % (spider.name, self.wuma_count))
            self.wb.save('C:\\Users\\lushe\\Desktop\\无码.xlsx')
            return item
        if spider.name == 'youma':
            line = [item['number'],item['title'],item['pdate'],item['detail_link'],item['cover'],item['fenlei'],item['actor'],item['maglinks']]
            self.ws.append(line)
            self.youma_count += 1
            print('已下载 %s, %d:' % (spider.name, self.youma_count))
            self.wb.save('C:\\Users\\lushe\\Desktop\\有码.xlsx')
            return item
        if spider.name == 'oumei':
            line = [item['number'],item['title'],item['pdate'],item['detail_link'],item['cover'],item['fenlei'],item['actor'],item['maglinks']]
            self.ws.append(line)
            self.oumei_count += 1
            print('已下载 %s, %d:' % (spider.name, self.oumei_count))
            self.wb.save('C:\\Users\\lushe\\Desktop\\欧美.xlsx')
            return item
        # if spider.name == 'youma':
        #     self.ws.append(line)
        #     YoumaSpider.count += 1
        #     print('已下载 %s, %d:' % (spider.name, YoumaSpider.count))
        #     spider.log(Warning, '已下载 %s, %d:' % (spider.name, YoumaSpider.count))
        #     return item
        # elif spider.name == 'wuma':
        #     self.ws.append(line)
        #     WumaSpider.count += 1
        #     print('已下载 %s, %d:' % (spider.name, WumaSpider.count))
        #     spider.log(Warning, '已下载 %s, %d:' % (spider.name, WumaSpider.count))
        #     return item
        # elif spider.name == 'oumei':
        #     self.ws.append(line)
        #     OumeiSpider.count += 1
        #     print('已下载 %s, %d:' % (spider.name, OumeiSpider.count))
        #     spider.log(Warning, '已下载 %s, %d:' % (spider.name, OumeiSpider.count))
        #     return item

    # def close_spider(self, spider):
    #     '''
    #     save lines to excel
    #     :return:
    #     '''
    #     print('ExcelPipline info: %s items size: %d' % spider.name, WumaSpider.count)
    #     self.wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
        # if spider.name == 'youma':
        #     print('ExcelPipline info: %s items size: %d' % spider.name, YoumaSpider.count)
        #     file_name = 'C:\\Users\\lushe\\Desktop\\有码.xlsx'
        #     self.wb.save(file_name)
        # elif spider.name == 'wuma':
        #     print('ExcelPipline info: %s items size: %d' % spider.name, WumaSpider.count)
        #     file_name = 'C:\\Users\\lushe\\Desktop\\无码.xlsx'
        #     self.wb.save(file_name)
        # elif spider.name == 'oumei':
        #     print('ExcelPipline info: %s items size: %d' % spider.name, OumeiSpider.count)
        #     file_name = 'C:\\Users\\lushe\\Desktop\\欧美.xlsx'
        #     self.wb.save(file_name)
