#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng


import time
import requests
from openpyxl import load_workbook
import openpyxl

from lxml import etree


headers = {
    r'cookie': '__cfduid=d761cee020cec276f507132139addc7421545555734; __guid=35400359.4113806376697964000.1545555738264.177; HstCfa4127494=1545555756697; HstCmu4127494=1545555756697; PHPSESSID=6ln21mg1n6nrhc3m5t9q9eqt04; starinfo=glyphicon%20glyphicon-plus; 4fJN_2132_saltkey=HJJ2dBHW; 4fJN_2132_lastvisit=1546156644; 4fJN_2132_sid=TRLVHQ; 4fJN_2132_lastact=1546160245%09misc.php%09seccode; 4fJN_2132_seccode=7942.ba32df629914d0e609; HstCnv4127494=3; HstCns4127494=7; monitor_count=136; HstCla4127494=1546171197997; HstPn4127494=74; HstPt4127494=137; existmag=all',
}

#首页
fenlei = {'无码': 'https://www.xwb2uooia5o59avwxb.com/uncensored/page/',
          '欧美': 'https://www.javbus.work/page/',
          '有码': 'https://www.xwb2uooia5o59avwxb.com/page/'}


# wb = load_workbook('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
wb = openpyxl.Workbook('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')


for leibie in fenlei.keys():
    #建立分类数据表
    wb.create_sheet(leibie)
    wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
    wb = load_workbook('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')
    # 获得所有sheet的名称
    print(wb.get_sheet_names())
    # 根据sheet名字获得sheet
    a_sheet = wb.get_sheet_by_name(leibie)
    a_sheet['A1'] = '番号'
    a_sheet['B1'] = '片名'
    a_sheet['C1'] = '发行时间'
    a_sheet['D1'] = '封面'
    a_sheet['E1'] = '详细页地址'
    wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')

    #最大查询到400页，可以修改为更大值，目前全站约300多页
    for n in range(1, 400):
        req = requests.get(fenlei[leibie]+str(n), headers=headers)
        # time.sleep(1)
        html = req.content.decode('utf-8', 'ignore')
        # print(html)
        page = etree.HTML(html)
        # print(page)
        print('开始下载第%d页' % n)
        #少数番号为空，不做下载，跳过
        if page.xpath("//*[@id='waterfall']/div[1]/a/div[2]/span/date[1]/text()") == []:
            break
        else:
            #获取页面中电影的信息
            for i in range(1, 31):
                if page.xpath("//*[@id='waterfall']/div[%d]/a/div[2]/span/date[1]/text()" % i) == []:
                    break
                else:
                    # print('下载第%d页第%d条数据' % n, i)
                    # 详细页链接
                    detail_link = page.xpath("//*[@id='waterfall']/div[%d]/a/@href" % i)
                    # print(detail_link)
                    # 封面图链接
                    cover_link = page.xpath("//*[@id='waterfall']//div[%d]/a/div[1]/img/@src" % i)
                    # print(cover_link)
                    # 片名
                    title = page.xpath("//*[@id='waterfall']//div[%d]/a/div[1]/img/@title" % i)
                    # print(title)
                    # 番号
                    item = page.xpath("//*[@id='waterfall']/div[%d]/a/div[2]/span/date[1]/text()" % i)
                    # print(item)
                    # 发行日期
                    pdate = page.xpath("//*[@id='waterfall']/div[%d]/a/div[2]/span/date[2]/text()" % i)
                    # print(pdate)
                    #番号+片名+发行日期+封面图链接+详细页链接
                    result = item + title + pdate + cover_link + detail_link
                    print('第%d页第%d条数据' % (n, i), result)

                    # print(a_sheet.max_row)
                    write_row = a_sheet.max_row + 1
                    a_sheet['A%d' % write_row] = result[0]
                    a_sheet['B%d' % write_row] = result[1]
                    a_sheet['C%d' % write_row] = result[2]
                    a_sheet['D%d' % write_row] = result[3]
                    a_sheet['E%d' % write_row] = result[4]
                    print('写入第%d条数据' % (write_row-1), result)

        wb.save('C:\\Users\\lushe\\Desktop\\新建工作表.xlsx')